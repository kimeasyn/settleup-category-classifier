import glob
import csv
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

CATEGORIES = ["식비", "교통", "숙박", "관광", "쇼핑", "기타"]

# data/raw/ 디렉토리의 원본 csv 파일 모두 읽기
data_dir = os.path.dirname(os.path.abspath(__file__))
raw_dir = os.path.join(data_dir, "raw")
csv_files = sorted(glob.glob(os.path.join(raw_dir, "*.csv")))

rows = []
for filepath in csv_files:
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader, None)  # 헤더 스킵
        for row in reader:
            if len(row) >= 2 and row[0].strip():
                rows.append([row[0].strip(), row[1].strip()])

print(f"CSV 파일 {len(csv_files)}개에서 총 {len(rows)}건 로드")

# xlsx 생성
wb = Workbook()
ws = wb.active
ws.title = "학습데이터"

# 헤더
ws.append(["description", "category"])
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")

# 데이터
for row in rows:
    ws.append(row)

# B열 드롭다운
dv = DataValidation(
    type="list",
    formula1='"' + ",".join(CATEGORIES) + '"',
    allow_blank=False,
)
dv.error = "목록에서 선택해주세요"
dv.errorTitle = "카테고리 오류"
dv.prompt = "카테고리를 선택하세요"
dv.promptTitle = "카테고리"
ws.add_data_validation(dv)
dv.add(f"B2:B{len(rows) + 1}")

# 열 너비
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 15

output_path = os.path.join(data_dir, "result.xlsx")
wb.save(output_path)
print(f"저장 완료: {output_path} ({len(rows)}건)")
